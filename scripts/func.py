# -*- coding: utf-8 -*-
#
# Copyright (C) 2022 Ing <https://github.com/wjz304>
#
# This is free software, licensed under the MIT License.
# See /LICENSE for more information.
#

import os, re, sys, glob, json, yaml, click, shutil, tarfile, kmodule, requests, urllib3
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry  # type: ignore
from openpyxl import Workbook

@click.group()
def cli():
    """
    The CLI is a commands to RR.
    """
    pass


@cli.command()
@click.option("-w", "--workpath", type=str, required=True, help="The workpath of RR.")
@click.option("-j", "--jsonpath", type=str, required=True, help="The output path of jsonfile.")
@click.option("-x", "--xlsxpath", type=str, required=False, help="The output path of xlsxfile.")
def getmodels(workpath, jsonpath, xlsxpath):
    models = {}
    platforms_yml = os.path.join(workpath, "opt", "rr", "platforms.yml")
    with open(platforms_yml, "r") as f:
        P_data = yaml.safe_load(f)
        P_platforms = P_data.get("platforms", [])
        for P in P_platforms:
            productvers = {}
            for V in P_platforms[P]["productvers"]:
                kpre = P_platforms[P]["productvers"][V].get("kpre", "")
                kver = P_platforms[P]["productvers"][V].get("kver", "")
                productvers[V] = f"{kpre}-{kver}" if kpre else kver
            models[P] = {"productvers": productvers, "models": []}

    adapter = HTTPAdapter(max_retries=Retry(total=3, backoff_factor=1, status_forcelist=[500, 502, 503, 504]))
    session = requests.Session()
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    try:
        url = "http://update7.synology.com/autoupdate/genRSS.php?include_beta=1"
        #url = "https://update7.synology.com/autoupdate/genRSS.php?include_beta=1"

        req = session.get(url, timeout=10, verify=False)
        req.encoding = "utf-8"
        p = re.compile(r"<mUnique>(.*?)</mUnique>.*?<mLink>(.*?)</mLink>", re.MULTILINE | re.DOTALL)
        data = p.findall(req.text)
    except Exception as e:
        click.echo(f"Error: {e}")
        return

    for item in data:
        if not "DSM" in item[1]:
            continue
        arch = item[0].split("_")[1]
        name = item[1].split("/")[-1].split("_")[1].replace("%2B", "+")
        if arch not in models:
            continue
        if name in (A for B in models for A in models[B]["models"]):
            continue
        models[arch]["models"].append(name)

    if jsonpath:
        with open(jsonpath, "w") as f:
            json.dump(models, f, indent=4, ensure_ascii=False)
    if xlsxpath:
        wb = Workbook()
        ws = wb.active
        ws.append(["platform", "productvers", "Model"])
        for k, v in models.items():
            ws.append([k, str(v["productvers"]), str(v["models"])])
        wb.save(xlsxpath)


@cli.command()
@click.option("-w", "--workpath", type=str, required=True, help="The workpath of RR.")
@click.option("-j", "--jsonpath", type=str, required=True, help="The output path of jsonfile.")
@click.option("-x", "--xlsxpath", type=str, required=False, help="The output path of xlsxfile.")
def getpats(workpath, jsonpath, xlsxpath):
    def __fullversion(ver):
        arr = ver.split('-')
        a, b, c = (arr[0].split('.') + ['0', '0', '0'])[:3]
        d = arr[1] if len(arr) > 1 else '00000'
        e = arr[2] if len(arr) > 2 else '0'
        return f'{a}.{b}.{c}-{d}-{e}'

    platforms_yml = os.path.join(workpath, "opt", "rr", "platforms.yml")
    with open(platforms_yml, "r") as f:
        data = yaml.safe_load(f)
        platforms = data.get("platforms", [])

    adapter = HTTPAdapter(max_retries=Retry(total=3, backoff_factor=1, status_forcelist=[500, 502, 503, 504]))
    session = requests.Session()
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    try:
        url = "http://update7.synology.com/autoupdate/genRSS.php?include_beta=1"
        #url = "https://update7.synology.com/autoupdate/genRSS.php?include_beta=1"

        req = session.get(url, timeout=10, verify=False)
        req.encoding = "utf-8"
        p = re.compile(r"<mUnique>(.*?)</mUnique>.*?<mLink>(.*?)</mLink>", re.MULTILINE | re.DOTALL)
        data = p.findall(req.text)
    except Exception as e:
        click.echo(f"Error: {e}")
        return

    models = []
    for item in data:
        if not "DSM" in item[1]:
            continue
        arch = item[0].split("_")[1]
        name = item[1].split("/")[-1].split("_")[1].replace("%2B", "+")
        if arch not in platforms:
            continue
        if name in models:
            continue
        models.append(name)

    pats = {}
    for M in models:
        pats[M] = {}
        version = '7'
        urlInfo = "https://www.synology.com/api/support/findDownloadInfo?lang=en-us"
        urlSteps = "https://www.synology.com/api/support/findUpgradeSteps?"
        #urlInfo = "https://www.synology.cn/api/support/findDownloadInfo?lang=zh-cn"
        #urlSteps = "https://www.synology.cn/api/support/findUpgradeSteps?"

        major = f"&major={version.split('.')[0]}" if len(version.split('.')) > 0 else ""
        minor = f"&minor={version.split('.')[1]}" if len(version.split('.')) > 1 else ""
        try:
            req = session.get(f"{urlInfo}&product={M.replace('+', '%2B')}{major}{minor}", timeout=10, verify=False)
            req.encoding = "utf-8"
            data = json.loads(req.text)
        except Exception as e:
            click.echo(f"Error: {e}")
            continue

        build_ver = data['info']['system']['detail'][0]['items'][0]['build_ver']
        build_num = data['info']['system']['detail'][0]['items'][0]['build_num']
        buildnano = data['info']['system']['detail'][0]['items'][0]['nano']
        V = __fullversion(f"{build_ver}-{build_num}-{buildnano}")
        if V not in pats[M]:
            pats[M][V] = {
                'url': data['info']['system']['detail'][0]['items'][0]['files'][0]['url'].split('?')[0],
                'sum': data['info']['system']['detail'][0]['items'][0]['files'][0]['checksum']
            }

        from_ver = min(I['build'] for I in data['info']['pubVers'])

        for I in data['info']['productVers']:
            if not I['version'].startswith(version):
                continue
            if not major or not minor:
                majorTmp = f"&major={I['version'].split('.')[0]}" if len(I['version'].split('.')) > 0 else ""
                minorTmp = f"&minor={I['version'].split('.')[1]}" if len(I['version'].split('.')) > 1 else ""
                try:
                    reqTmp = session.get(f"{urlInfo}&product={M.replace('+', '%2B')}{majorTmp}{minorTmp}", timeout=10, verify=False)
                    reqTmp.encoding = "utf-8"
                    dataTmp = json.loads(reqTmp.text)
                except Exception as e:
                    click.echo(f"Error: {e}")
                    continue

                build_ver = dataTmp['info']['system']['detail'][0]['items'][0]['build_ver']
                build_num = dataTmp['info']['system']['detail'][0]['items'][0]['build_num']
                buildnano = dataTmp['info']['system']['detail'][0]['items'][0]['nano']
                V = __fullversion(f"{build_ver}-{build_num}-{buildnano}")
                if V not in pats[M]:
                    pats[M][V] = {
                        'url': dataTmp['info']['system']['detail'][0]['items'][0]['files'][0]['url'].split('?')[0],
                        'sum': dataTmp['info']['system']['detail'][0]['items'][0]['files'][0]['checksum']
                    }

            for J in I['versions']:
                to_ver = J['build']
                try:
                    reqSteps = session.get(f"{urlSteps}&product={M.replace('+', '%2B')}&from_ver={from_ver}&to_ver={to_ver}", timeout=10, verify=False)
                    if reqSteps.status_code != 200:
                        continue
                    reqSteps.encoding = "utf-8"
                    dataSteps = json.loads(reqSteps.text)
                except Exception as e:
                    click.echo(f"Error: {e}")
                    continue

                for S in dataSteps['upgrade_steps']:
                    if not S.get('full_patch') or not S['build_ver'].startswith(version):
                        continue
                    V = __fullversion(f"{S['build_ver']}-{S['build_num']}-{S['nano']}")
                    if V not in pats[M]:
                        pats[M][V] = {
                            'url': S['files'][0]['url'].split('?')[0],
                            'sum': S['files'][0]['checksum']
                        }

    if jsonpath:
        with open(jsonpath, "w") as f:
            json.dump(pats, f, indent=4, ensure_ascii=False)
    if xlsxpath:
        wb = Workbook()
        ws = wb.active
        ws.append(["Model", "version", "url", "sum"])
        for k1, v1 in pats.items():
            for k2, v2 in v1.items():
                ws.append([k1, k2, v2["url"], v2["sum"]])
        wb.save(xlsxpath)


@cli.command()
@click.option("-w", "--workpath", type=str, required=True, help="The workpath of RR.")
@click.option("-j", "--jsonpath", type=str, required=True, help="The output path of jsonfile.")
@click.option("-x", "--xlsxpath", type=str, required=False, help="The output path of xlsxfile.")
def getaddons(workpath, jsonpath, xlsxpath):
    AS = glob.glob(os.path.join(workpath, "mnt", "p3", "addons", "*", "manifest.yml"))
    AS.sort()
    addons = {}
    for A in AS:
        with open(A, "r") as file:
            A_data = yaml.safe_load(file)
            A_name = A_data.get("name", "")
            A_system = A_data.get("system", False)
            A_description = A_data.get("description", {"en_US": "Unknown", "zh_CN": "Unknown"})
            addons[A_name] = {"system": A_system, "description": A_description}
    if jsonpath:
        with open(jsonpath, "w") as f:
            json.dump(addons, f, indent=4, ensure_ascii=False)
    if xlsxpath:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "system", "en_US", "zh_CN"])
        for k1, v1 in addons.items():
            ws.append([k1, v1.get("system", False), v1.get("description").get("en_US", ""), v1.get("description").get("zh_CN", "")])
        wb.save(xlsxpath)


@cli.command()
@click.option("-w", "--workpath", type=str, required=True, help="The workpath of RR.")
@click.option("-j", "--jsonpath", type=str, required=True, help="The output path of jsonfile.")
@click.option("-x", "--xlsxpath", type=str, required=False, help="The output path of xlsxfile.")
def getmodules(workpath, jsonpath, xlsxpath):
    MS = glob.glob(os.path.join(workpath, "mnt", "p3", "modules", "*.tgz"))
    MS.sort()
    modules = {}
    TMP_PATH = "/tmp/modules"
    if os.path.exists(TMP_PATH):
        shutil.rmtree(TMP_PATH)
    for M in MS:
        M_name = os.path.splitext(os.path.basename(M))[0]
        M_modules = {}
        os.makedirs(TMP_PATH)
        with tarfile.open(M, "r") as tar:
            tar.extractall(TMP_PATH)
        KS = glob.glob(os.path.join(TMP_PATH, "*.ko"))
        KS.sort()
        for K in KS:
            K_name = os.path.splitext(os.path.basename(K))[0]
            K_info = kmodule.modinfo(K, basedir=os.path.dirname(K), kernel=None)[0]
            K_description = K_info.get("description", "")
            K_depends = K_info.get("depends", "")
            M_modules[K_name] = {"description": K_description, "depends": K_depends}
        modules[M_name] = M_modules
        if os.path.exists(TMP_PATH):
            shutil.rmtree(TMP_PATH)
    if jsonpath:
        with open(jsonpath, "w") as file:
            json.dump(modules, file, indent=4, ensure_ascii=False)
    if xlsxpath:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Arch", "description", "depends"])
        for k1, v1 in modules.items():
            for k2, v2 in v1.items():
                ws.append([k2, k1, v2["description"], v2["depends"]])
        wb.save(xlsxpath)


if __name__ == "__main__":
    cli()
